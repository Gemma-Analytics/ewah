import io
import requests
from msal import ConfidentialClientApplication
from openpyxl import load_workbook
from urllib.parse import quote

from ewah.hooks.base import EWAHBaseHook


class EWAHSharepointHook(EWAHBaseHook):
    GRAPH_API_BASE = "https://graph.microsoft.com"

    conn_name_attr: str = "ewah_sharepoint_conn_id"
    default_conn_name: str = "ewah_sharepoint_default"
    conn_type: str = "ewah_sharepoint"
    hook_name: str = "EWAH Microsoft Sharepoint Connection"

    @staticmethod
    def get_ui_field_behaviour() -> dict:
        return {
            "hidden_fields": ["port", "extra", "host"],
            "relabeling": {
                "login": "Client ID",
                "password": "Client Secret",
                "schema": "Tenant ID",
            },
        }

    @staticmethod
    def get_connection_form_widgets() -> dict:
        """Returns connection widgets to add to connection form"""
        from wtforms import StringField
        from flask_appbuilder.fieldwidgets import BS3TextFieldWidget

        return {
            "extra__ewah_sharepoint__site_hostname": StringField(
                "Site Hostname (e.g. company.sharepoint.com)",
                widget=BS3TextFieldWidget(),
            ),
            "extra__ewah_sharepoint__site_path": StringField(
                "Site Path (e.g. sites/Finance)",
                widget=BS3TextFieldWidget(),
            ),
            "extra__ewah_sharepoint__drive_id": StringField(
                "Drive ID (optional, e.g. b!oS9JMAuH1...DdrP)",
                widget=BS3TextFieldWidget(),
            ),
        }

    def _authenticate_with_graph_api(self):
        """Authenticate with Microsoft Graph API using client credentials flow."""
        self.log.info("Authenticating with Microsoft Graph API...")

        # Get credentials from connection
        client_id = self.conn.login
        client_secret = self.conn.password
        tenant_id = self.conn.schema

        if not all([client_id, client_secret, tenant_id]):
            raise ValueError(
                "client_id, client_secret, and tenant_id must be configured in the connection"
            )

        authority = f"https://login.microsoftonline.com/{tenant_id}"
        scopes = ["https://graph.microsoft.com/.default"]

        try:
            app = ConfidentialClientApplication(
                client_id, authority=authority, client_credential=client_secret
            )

            result = app.acquire_token_for_client(scopes=scopes)

            if "access_token" in result:
                self.log.info("Successfully authenticated with Microsoft Graph API")
                return result["access_token"]
            else:
                error_msg = f"Authentication failed: {result.get('error')} - {result.get('error_description')}"
                self.log.error(error_msg)
                raise Exception(error_msg)
        except Exception as e:
            self.log.error(f"Authentication error: {str(e)}")
            raise

    def _get_site_id(self, access_token, site_hostname, site_path):
        """Get the SharePoint site ID from the site hostname and path."""
        self.log.info(f"Getting site ID for {site_hostname}/{site_path}")
        site_url = f"{self.GRAPH_API_BASE}/v1.0/sites/{site_hostname}:/{site_path}"
        site_resp = requests.get(
            site_url, headers={"Authorization": f"Bearer {access_token}"}
        )
        site_resp.raise_for_status()
        site_id = site_resp.json()["id"]
        self.log.info(f"Retrieved site ID: {site_id}")
        return site_id

    def _get_drive_id(self, access_token, site_id):
        """Get the default drive ID from the SharePoint site."""
        self.log.info(f"Getting drive ID for site {site_id}")
        drive_url = f"{self.GRAPH_API_BASE}/v1.0/sites/{site_id}/drives"
        drive_resp = requests.get(
            drive_url, headers={"Authorization": f"Bearer {access_token}"}
        )
        drive_resp.raise_for_status()
        drive_id = drive_resp.json()["value"][0]["id"]
        self.log.info(f"Retrieved drive ID: {drive_id}")
        return drive_id

    def get_data_from_excel(
        self,
        relative_file_path: str,
        worksheet_name: str,
        header_row: int,
        start_row: int,
        batch_size: int,
        site_hostname: str = None,
        site_path: str = None,
        drive_id: str = None,
    ):
        """
        Get data from Excel file in SharePoint using Microsoft Graph API.

        Args:
            relative_file_path: Path to the Excel file relative to the SharePoint site
            worksheet_name: Name of the worksheet to read
            header_row: Row number containing headers (1-based)
            start_row: Row number to start reading data from (1-based)
            batch_size: Number of rows to process in each batch
            site_hostname: SharePoint hostname (e.g., "company.sharepoint.com")
            site_path: SharePoint site path (e.g., "sites/MySite")
            drive_id: SharePoint drive ID
        """
        assert (
            isinstance(batch_size, int) and batch_size > 0
        ), f"type: {type(batch_size)}, value: {str(batch_size)}"
        assert (
            isinstance(header_row, int) and header_row > 0
        ), f"type: {type(header_row)}, value: {str(header_row)}"
        assert (
            isinstance(start_row, int) and start_row > 0
        ), f"type: {type(start_row)}, value: {str(start_row)}"
        assert start_row > header_row

        # Get site_hostname and site_path from connection if not provided
        if not site_hostname:
            site_hostname = self.conn.extra_dejson.get(
                "extra__ewah_sharepoint__site_hostname"
            )
        if not site_path:
            site_path = self.conn.extra_dejson.get("extra__ewah_sharepoint__site_path")

        if not drive_id:
            drive_id = self.conn.extra_dejson.get("extra__ewah_sharepoint__drive_id")

        if not site_hostname or not site_path:
            raise ValueError(
                "site_hostname and site_path must be provided either as parameters or in the connection"
            )

        # Authenticate with Microsoft Graph API
        access_token = self._authenticate_with_graph_api()

        # Get site ID
        site_id = self._get_site_id(access_token, site_hostname, site_path)

        # If drive_id is not provided, get it using the default method
        if not drive_id:
            drive_id = self._get_drive_id(access_token, site_id)

        # Download the file using Graph API
        self.log.info(f"Downloading file: {relative_file_path}")
        file_path_encoded = quote(relative_file_path.strip().strip("/"))
        download_url = f"{self.GRAPH_API_BASE}/v1.0/sites/{site_id}/drives/{drive_id}/root:/{file_path_encoded}:/content"

        file_resp = requests.get(
            download_url, headers={"Authorization": f"Bearer {access_token}"}
        )
        file_resp.raise_for_status()

        # Process the file content
        bytes_file_obj = io.BytesIO()
        bytes_file_obj.write(file_resp.content)
        bytes_file_obj.seek(0)

        try:
            workbook = load_workbook(bytes_file_obj, data_only=True)
            if worksheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{worksheet_name}' not found. Available worksheets: {workbook.sheetnames}"
                )
            ws = workbook[worksheet_name]

            # Get headers
            headers = {
                ws.cell(row=header_row, column=col).value: col
                for col in range(1, ws.max_column + 1)
                if ws.cell(row=header_row, column=col).value is not None
            }

            if not headers:
                self.log.warning(f"No headers found in row {header_row}")
                return

            self.log.info(f"Found {len(headers)} columns: {list(headers.keys())}")

            # Process data in batches
            while start_row <= ws.max_row:
                self.log.info(
                    f"Yielding batch of {batch_size} rows starting from row {start_row}."
                )
                batch_data = []
                for row in range(
                    start_row, min(start_row + batch_size, ws.max_row + 1)
                ):
                    row_data = {}
                    for k, v in headers.items():
                        cell_value = ws.cell(row=row, column=v).value
                        row_data[k] = cell_value
                    batch_data.append(row_data)

                if batch_data:
                    yield batch_data
                start_row += batch_size

        except Exception as e:
            self.log.error(f"Error processing Excel file: {str(e)}")
            raise
